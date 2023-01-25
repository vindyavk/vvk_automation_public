import re
import config
import pytest
import unittest
import logging
import subprocess
import os
import json
import ib_utils.ib_NIOS as ib_NIOS
import commands
import json, ast
import requests
from time import sleep as sleep
import pexpect
import paramiko
import time
import sys
import socket
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from paramiko import client
from ib_utils.start_stop_logs import log_action as log
from ib_utils.file_content_validation import log_validation as logv
import openpyxl
f = open("rfe-9986.log",'w')

def Average(lst):
    return sum(lst) / len(lst)

def create_xlsx():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet_title = sheet.title
    sheet.title = "WAPI Performance"
    c1=sheet.cell(row= 2 , column = 1)
    c1.value="SRV"
    c2=sheet.cell(row= 3 , column = 1)
    c2.value="CNAME"
    c3=sheet.cell(row= 4 , column = 1)
    c3.value="DNAME"
    c4=sheet.cell(row= 5 , column = 1)
    c4.value="MX"
    c5=sheet.cell(row= 6 , column = 1)
    c5.value="TXT"
    c6=sheet.cell(row= 7 , column = 1)
    c6.value="NAPTR"
    c7=sheet.cell(row= 8 , column = 1)
    c7.value="A"
    c8=sheet.cell(row= 9 , column = 1)
    c8.value="NS"
    c9=sheet.cell(row= 1 , column = 2)
    c9.value="Average time taken"
    c10=sheet.cell(row= 1 , column = 3)
    c10.value="Minimum time taken"

    li_of_srv=[]
    li_of_cname=[]
    li_of_dname=[]
    li_of_mx=[]
    li_of_txt=[]
    li_of_naptr=[]
    li_of_a=[]
    li_of_ns=[]

    for i in range(config.iterations):
        print config.iterations
       
        srv=Get_SRV_record(config.no)
        print srv
        li_of_srv.append(srv)

        cname=Get_CNAME_record(config.no)
        li_of_cname.append(cname)

        dname=Get_DNAME_record(config.no)
        li_of_dname.append(dname)

        mx=Get_MX_record(config.no)
        li_of_mx.append(mx)

        txt=Get_TXT_record(config.no)
        li_of_txt.append(txt)

        naptr=Get_NAPTR_record(config.no)
        li_of_naptr.append(naptr)

        a=Get_A_record(config.no)
        li_of_a.append(a)

        ns=Get_NS_record(config.no)
        li_of_ns.append(ns)
    print li_of_srv
    avg_of_srv=Average(li_of_srv)
    min_of_srv=min(li_of_srv)

    avg_of_cname=Average(li_of_cname)
    min_of_cname=min(li_of_cname)

    avg_of_dname=Average(li_of_dname)
    min_of_dname=min(li_of_dname)

    avg_of_mx=Average(li_of_mx)
    min_of_mx=min(li_of_mx)

    avg_of_txt=Average(li_of_txt)
    min_of_txt=min(li_of_txt)
    
    avg_of_naptr=Average(li_of_naptr)
    min_of_naptr=min(li_of_naptr)

    avg_of_a=Average(li_of_a)
    min_of_a=min(li_of_a)

    avg_of_ns=Average(li_of_ns)
    min_of_ns=min(li_of_ns)


    c11=sheet.cell(row= 2 , column = 2)
    c11.value=avg_of_srv
    c12=sheet.cell(row= 2 , column = 3)
    c12.value=min_of_srv

    c11=sheet.cell(row= 3 , column = 2)
    c11.value=avg_of_cname
    c12=sheet.cell(row= 3 , column = 3)
    c12.value=min_of_cname

    c11=sheet.cell(row= 4 , column = 2)
    c11.value=avg_of_dname
    c12=sheet.cell(row= 4 , column = 3)
    c12.value=min_of_dname

    c11=sheet.cell(row= 5 , column = 2)
    c11.value=avg_of_mx
    c12=sheet.cell(row= 5 , column = 3)
    c12.value=min_of_mx

    c11=sheet.cell(row= 6 , column = 2)
    c11.value=avg_of_txt
    c12=sheet.cell(row= 6 , column = 3)
    c12.value=min_of_txt

    c11=sheet.cell(row= 7 , column = 2)
    c11.value=avg_of_naptr
    c12=sheet.cell(row= 7 , column = 3)
    c12.value=min_of_naptr

    c11=sheet.cell(row= 8 , column = 2)
    c11.value=avg_of_a
    c12=sheet.cell(row= 8 , column = 3)
    c12.value=min_of_a

    c11=sheet.cell(row= 9 , column = 2)
    c11.value=avg_of_ns
    c12=sheet.cell(row= 9 , column = 3)
    c12.value=min_of_ns
    
    wb.save("9986_wapi_performance.xlsx")

def Get_A_record(no):
    total_time=0
    for i in range(no):
        #total_time=0
        start_time = time.time()
        response = ib_NIOS.wapi_request('GET', object_type="record:a?name=zzwq56gkn0pag.k3a95ds.zone1")
        end_time = time.time()
        if response=="[]":
            assert False
        time1=end_time-start_time
        total_time=total_time+time1
        print>>f,("Time taken to GET A "+str(i)+"th record",time1)
        print>>f,(response)
    print("Total time took to Get %d A records is %f seconds"%(no,total_time))
    return(total_time)


def Get_SRV_record(no):
    total_time=0
    for i in range(no):
        start_time = time.time()
        response = ib_NIOS.wapi_request('GET', object_type="record:srv?name=99986gyjbkedkwzqe3p.wb4eejy6rbuv6dk9kbwingizowxwoe.1zone1")
        end_time = time.time()
        if response=="[]":
            assert False
        time1=end_time-start_time
        total_time=total_time+time1
        print>>f,("Time taken to GET SRV "+str(i)+"th record",time1)
        print>>f,(response)

    print("Total time took to Get %d SRV records is %f seconds for SRV"%(no,total_time))
    return(total_time)

def Get_CNAME_record(no):
    total_time=0
    for i in range(no):
        start_time = time.time()
        response = ib_NIOS.wapi_request('GET', object_type="record:cname?name=999882npuratpt.w9.1zone1" )
        end_time = time.time()
        if response=="[]":
            assert False
        time1=end_time-start_time
        total_time=total_time+time1
        print>>f,("Time taken to get CNAME "+str(i)+"th record",time1)
        print>>f,(response)
    print("Total time took to get %d CNAME records is %f seconds for CNAME"%(no,total_time))
    return(total_time)

def Get_DNAME_record(no):
    total_time=0
    for i in range(no):
        start_time = time.time()
        response = ib_NIOS.wapi_request('GET', object_type="record:dname?name=99986xi6r.twdzv84z5rjc0.1zone1")
        end_time = time.time()
        if response=="[]":
            assert False
        time1=end_time-start_time
        total_time=total_time+time1
        print>>f,("Time taken to Get Dname "+str(i)+"th record",time1)
        print>>f,(response)
    print("Total time took to Get %d DNAME records is %f seconds for DNAME"%(no,total_time))
    return(total_time)

def Get_MX_record(no):
    total_time=0
    for i in range(no):
        start_time = time.time()
        response = ib_NIOS.wapi_request('GET', object_type="record:mx?name=99987envhqo27ml.fmiacxgo2dmmzr34kconze-wbq.1zone1")
        end_time = time.time()
        if response=="[]":
            assert False
        time1=end_time-start_time
        total_time=total_time+time1
        print>>f,("Time taken to Get MX "+str(i)+"th record",time1)
        print>>f,(response)
    print("Total time took to Get %d MX records is %f seconds for MX"%(no,total_time))
    return(total_time)

def Get_NAPTR_record(no):
    total_time=0
    for i in range(no):
        start_time = time.time()
        response = ib_NIOS.wapi_request('GET', object_type="record:naptr?name=99985mlip6e1cyqpg3kyre76pise.wlizl3ozopv8ejyt73gomxfblenwn0.1zone1")
        end_time = time.time()
        if response=="[]":
            assert False
        time1=end_time-start_time
        total_time=total_time+time1
        print>>f,("Time taken to Get NAPTR "+str(i)+"th record",time1)
        print>>f,(response)
    print("Total time took to Get %d NAPTR records is %f seconds for NAPTR"%(no,total_time))
    return(total_time)

def Get_TXT_record(no):
    total_time=0
    for i in range(no):
        start_time = time.time()
        response = ib_NIOS.wapi_request('GET', object_type="record:txt?name=99989wdfhbftrm.9i-gpsfmeby5knh.1zone1")
        end_time = time.time()
        if response=="[]":
            assert False
        time1=end_time-start_time
        total_time=total_time+time1
        print>>f,("Time taken to Get TXT "+str(i)+"th record",time1)
        print>>f,(response)
    print("Total time took to Get %d TXT records is %f seconds for TXT"%(no,total_time))
    return(total_time)


def Get_NS_record(no):
    total_time=0
    for i in range(no):
        start_time = time.time()
        response = ib_NIOS.wapi_request('GET', object_type="record:ns?name=99987czisrxq4sabvf8xhhrbqm.xefxls4hlawkfwcc3m5.1zone1")
        end_time = time.time()
        if response=="[]":
            assert False
        time1=end_time-start_time
        total_time=total_time+time1
        print>>f,("Time taken to Get NS "+str(i)+"th record",time1)
        print>>f,(response)
    print("Total time took to Get %d NS records is %f seconds for NS"%(no,total_time))
    return(total_time)

class Network(unittest.TestCase):
    @pytest.mark.run(order=1)
    def test_001_GRID_RESTORE(self):
             logging.info("Grid Restore")
             response = ib_NIOS.wapi_request('POST', object_type="fileop?_function=uploadinit")
             print response
             res = json.loads(response)
             URL=res['url']
             token1=res['token']
             print("URL is : %s", URL)
             print("Token is %s",token1)
             infoblox_log_validation ='curl -k -u admin:infoblox -H content_type="content-typemultipart-formdata" ' + str(URL) +' -F file=@database.bak'
             print infoblox_log_validation
             out2 = commands.getoutput(infoblox_log_validation)
             print ("out2$$$$$$",out2)
             data2={"mode":"FORCED","nios_data":True,"token":token1,"keep_grid_ip":True}
             print ("&*&*&*&*&*&*",data2)
             response2 = ib_NIOS.wapi_request('POST', object_type="fileop?_function=restoredatabase",fields=json.dumps(data2))
             sleep(500)
             logging.info("Validate Syslog afer perform queries")
             infoblox_log_validation = 'ssh -o StrictHostKeyChecking=no -o BatchMode=yes -o UserKnownHostsFile=/dev/null root@' + str( config.grid_vip) + ' " tail -2000 /infoblox/var/infoblox.log "'
             out1 = commands.getoutput(infoblox_log_validation)
             print out1
             logging.info(out1)
             assert re.search(r'restore_node complete',out1)
             sleep(50)
             read  = re.search(r'201',response)
             for read in  response:
                 assert True
             logging.info("Test Case 72 Execution Completed")

    @pytest.mark.run(order=2)
    def test_002_Start_DNS_Service(self):
        logging.info("Start DNS Service")
        get_ref = ib_NIOS.wapi_request('GET', object_type="member:dns")
        logging.info(get_ref)
        res = json.loads(get_ref)
        for i in res:
            ref=i["_ref"]
            logging.info("Modify a enable_dns")
            data = {"enable_dns": True}
            response = ib_NIOS.wapi_request('PUT', ref=ref, fields=json.dumps(data))
            sleep(5)
            logging.info(response)
            read  = re.search(r'200',response)
            for read in  response:
                assert True
        logging.info("Test Case Execution Completed")

    @pytest.mark.run(order=3)
    def test_003_Validate_DNS_service_is_Enabled(self):
        logging.info("Validate DNS Service is enabled")
        get_tacacsplus = ib_NIOS.wapi_request('GET', object_type="member:dns",params="?_return_fields=enable_dns")
        logging.info(get_tacacsplus)
        res = json.loads(get_tacacsplus)
        logging.info(res)
        for i in res:
            if i["enable_dns"] == True:
                logging.info("Test Case Execution Passed")
                assert True
            else:
                logging.info("Test Case Execution Failed")
                assert False

    @pytest.mark.run(order=4)
    def test_004_get_the_time_for_each_objects(self):
        logging.info("get the time for each objects")
        create_xlsx()
	assert True
        
    @pytest.mark.run(order=5)
    def test_005_send_an_email_with_the_result(self):
          data = MIMEMultipart()
          sender = 'jenkins@infoblox.com'
          #receivers = ['acm@infoblox.com']
          receivers = config.mail_list
          data['Subject'] = "9986-rfe-results"
          body = "Please Analyse the 9986 rfe results"
          data.attach(MIMEText(body, 'plain'))
          filename = ["9986_wapi_performance.xlsx","rfe-9986.log"]
          for i in filename:
             #file_name="/import/qaddi/API_Automation_20_09/WAPI_PyTest/suites/whitelist/"+str(i)
             file_name=os.getcwd()+'/'+str(i)
             attachment = open(file_name,"rb")
             #print(attachment.read())
             p = MIMEBase('application', 'octet-stream')
             p.set_payload((attachment).read())
             encoders.encode_base64(p)
             p.add_header('Content-Disposition', "attachment; filename= %s" % i)
             data.attach(p)
             message = data.as_string()
          try:
             smtpObj = smtplib.SMTP('localhost')
             smtpObj.sendmail(sender, receivers, message)
             print "Successfully sent email"
             assert True
          except smtplib.SMTPException:
             print "Error: unable to send email"
             assert False

    @pytest.mark.run(order=6)
    def test_006_delete_the_report_file(self):
        logging.info("delete the report file")
        command1=os.system("rm -rf 9986_wapi_performance.xlsx rfe-9986.log")
	


